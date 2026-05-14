"""
第2工場 生産機械抽出ツール（全部盛り完成版）
"""

import sys
import threading
import multiprocessing
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from datetime import datetime
import os

# ─── ログ関数 ─────────────────────────
def log_all(msg):
    t = datetime.now().strftime("%H:%M:%S")
    line = f"[{t}] {msg}"

    print(line)

    try:
        with open("debug_log.txt", "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except:
        pass

    try:
        app._log(msg)
    except:
        pass


# ─── openpyxl ─────────────────────────
try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    log_all("openpyxl 未インストール")
    sys.exit(0)


# ─── 設定 ─────────────────────────
TARGET_GROUP = "第2工場"

COL = {
    "OrderNo": 3,
    "機種グループ": 9,
    "機種": 11,
    "号機": 13,
}


# ─── ロジック ─────────────────────────
def _find_data_start(ws):
    log_all("① ヘッダー検索")
    for r in range(1, 200):
        for c in range(1, 30):
            val = ws.cell(r, c).value
            if val and str(val).strip() == "OrderNo":
                return r + 1
    raise ValueError("OrderNoが見つからない")


def _extract_records(ws, start):
    log_all("② 抽出開始")

    records = []
    seen = set()

    for r in range(start, ws.max_row + 1):
        if ws.cell(r, COL["機種グループ"]).value != TARGET_GROUP:
            continue

        order = ws.cell(r, COL["OrderNo"]).value

        if order in seen:
            continue
        seen.add(order)

        records.append({
            "OrderNo": order,
            "機種": ws.cell(r, COL["機種"]).value,
            "号機": ws.cell(r, COL["号機"]).value,
        })

    return records


def run_extraction(input_path, output_path):
    log_all("=== 処理開始 ===")
    log_all(f"入力: {input_path}")
    log_all(f"出力: {output_path}")

    log_all("ファイル読み込み")
    wb = openpyxl.load_workbook(input_path)

    if "main" not in wb.sheetnames:
        raise ValueError("mainシートなし")

    ws = wb["main"]

    start = _find_data_start(ws)
    records = _extract_records(ws, start)

    log_all(f"件数: {len(records)}")

    if not records:
        raise ValueError("データがありません")

    wb_out = Workbook()
    ws2 = wb_out.active
    ws2.title = "抽出結果"

    for i, r in enumerate(records, start=1):
        ws2.cell(i, 1, r["OrderNo"])
        ws2.cell(i, 2, r["機種"])
        ws2.cell(i, 3, r["号機"])

    log_all("保存")
    wb_out.save(output_path)

    # ✅ Excelで開く
    try:
        os.startfile(output_path)
    except:
        import subprocess
        subprocess.Popen(["start", "", output_path], shell=True)

    log_all("完了")

    return records


# ─── GUI ─────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        global app
        app = self

        self.title("第2工場抽出ツール")
        self.geometry("520x420")

        self.input = tk.StringVar()
        self.output = tk.StringVar()

        tk.Label(self, text="入力ファイル").pack()
        tk.Entry(self, textvariable=self.input, width=60).pack()
        tk.Button(self, text="参照", command=self.browse_in).pack()

        tk.Label(self, text="出力ファイル").pack()
        tk.Entry(self, textvariable=self.output, width=60).pack()
        tk.Button(self, text="参照", command=self.browse_out).pack()

        tk.Button(self, text="実行", command=self.run, bg="#4CAF50", fg="white").pack(pady=10)

        self.log_text = tk.Text(self, height=12)
        self.log_text.pack(fill="both", expand=True)

    def browse_in(self):
        f = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xlsm")]
        )
        if not f:
            return

        self.input.set(f)

        # ✅ 出力自動生成
        p = Path(f)
        today = datetime.now().strftime("%Y%m%d")
        auto = p.parent / f"{p.stem}_抽出_{today}.xlsx"
        self.output.set(str(auto))

    def browse_out(self):
        f = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if f:
            self.output.set(f)

    def _log(self, msg):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")

    def run(self):
        inp = self.input.get().strip()
        out = self.output.get().strip()

        if not inp:
            messagebox.showerror("エラー", "入力ファイルを選択してください")
            return

        if not out:
            messagebox.showerror("エラー", "出力先を指定してください")
            return

        def worker():
            try:
                log_all("スレッド開始")
                run_extraction(inp, out)
                log_all("成功")
            except Exception as e:
                import traceback
                err = traceback.format_exc()
                log_all("エラー発生")
                log_all(err)
                messagebox.showerror("エラー", str(e))

        threading.Thread(target=worker, daemon=True).start()


# ─── 起動 ─────────────────────────
if __name__ == "__main__":
    multiprocessing.freeze_support()

    log_all("アプリ起動")

    try:
        App().mainloop()
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        input("Enterで終了")
